"""Reconciliation engine — automatic, forced, and exclusion logic."""
from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from decimal import Decimal
from typing import List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.db.locks import LockKey, try_advisory_lock
from app.models.exclusion import Exclusion
from app.models.flow import Flow
from app.models.match_group import MatchGroup, MatchMode
from app.models.reconciliation_entry import EntryStatus
from app.models.reconciliation_run import ReconciliationRun
from app.repositories.exclusion_repository import exclusion_repository
from app.repositories.match_group_repository import match_group_repository
from app.repositories.movement_split_repository import movement_split_repository
from app.repositories.reconciliation_entry_repository import reconciliation_entry_repository
from app.repositories.reconciliation_run_repository import reconciliation_run_repository
from app.services.payment_status_service import payment_status_service

logger = logging.getLogger(__name__)


class ReconciliationAlreadyRunning(ValueError):
    """Another process holds the reconciliation_entry writer lock — a run_auto
    or an émargement sweep is already in progress.

    Subclasses ValueError to follow the repo's typed business-error convention,
    but the endpoints map it to 200/skipped rather than a 4xx: nothing is wrong,
    someone else is simply doing the work.
    """


class ReconciliationService:
    # --------------------------------------------------------------
    # List / query helpers
    # --------------------------------------------------------------
    def list_runs(self, db: Session, *, skip: int = 0, limit: int = 100) -> list:
        return reconciliation_run_repository.list(db, skip=skip, limit=limit)

    def list_entries_filtered(
        self,
        db: Session,
        *,
        ids=None,
        flow_id=None,
        status=None,
        reco_id=None,
        amount_min=None,
        amount_max=None,
        payment_statuses=None,
        payment_timestamp_from=None,
        payment_timestamp_to=None,
        account=None,
        date_from=None,
        date_to=None,
        search=None,
        skip: int = 0,
        limit: int = 100,
        with_payment_statuses: bool = True,
    ):
        filter_kwargs = dict(
            ids=ids,
            flow_id=flow_id,
            status=status,
            reco_id=reco_id,
            amount_min=amount_min,
            amount_max=amount_max,
            payment_statuses=payment_statuses,
            payment_timestamp_from=payment_timestamp_from,
            payment_timestamp_to=payment_timestamp_to,
            account=account,
            date_from=date_from,
            date_to=date_to,
            search=search,
        )
        items = reconciliation_entry_repository.list_filtered(db, **filter_kwargs, skip=skip, limit=limit)
        total_count = reconciliation_entry_repository.count_filtered(db, **filter_kwargs)
        if with_payment_statuses:
            # Per-reco payment aggregate ("3 ACC · 2 PDNG" + the timestamp span
            # in the operational view) — one query for the page. Payments are
            # keyed by reco_id, so every entry of a reconciliation group shows the
            # same set; entries without a reco_id show nothing.
            reco_ids = {e.reco_id for e in items if e.reco_id}
            agg = payment_status_service.aggregate_for_reco_ids(
                db, reco_ids=list(reco_ids)
            )
            for entry in items:
                summary = agg.get(entry.reco_id)
                setattr(entry, "payment_statuses", summary.statuses if summary else None)
                setattr(entry, "payment_timestamp_min", summary.payment_timestamp_min if summary else None)
                setattr(entry, "payment_timestamp_max", summary.payment_timestamp_max if summary else None)
        return items, total_count

    def distinct_payment_statuses(self, db: Session) -> List[str]:
        """Distinct payment-status values (for the operational filter options)."""
        return payment_status_service.distinct_statuses(db)

    def export_entries_excel(self, db: Session, *, filters) -> io.BytesIO:
        """Generate an Excel file with every entry matching ``filters``.

        Mirrors the archive export (``time_machine_service.export_snapshot_excel``):
        one styled "Entries" sheet, fetched with no pagination so the file holds
        all matching rows — not just the 200-row batches the operational view
        lazy-loads. Internal identifiers (db id, numeric flow_id, run / match ids)
        are intentionally omitted; the flow is shown by its human-readable code.
        """
        items, _ = self.list_entries_filtered(
            db,
            flow_id=filters.flow_id,
            status=filters.status,
            reco_id=filters.reco_id,
            amount_min=filters.amount_min,
            amount_max=filters.amount_max,
            payment_statuses=filters.payment_statuses,
            payment_timestamp_from=filters.payment_timestamp_from,
            payment_timestamp_to=filters.payment_timestamp_to,
            account=filters.account,
            date_from=filters.date_from,
            date_to=filters.date_to,
            search=filters.search,
            skip=0,
            limit=100_000,
            with_payment_statuses=True,  # export mirrors the table's "Payments" column
        )

        # flow_id → code so the export shows the same label as the UI table.
        flow_codes = {f.id: f.code for f in db.query(Flow).all()}

        wb = Workbook()
        ws = wb.active
        ws.title = "Entries"

        header_fill = PatternFill(start_color="2B2D42", end_color="2B2D42", fill_type="solid")
        header_text = Font(bold=True, color="FFFFFF")

        def _enum(v):
            return v.value if v is not None else None

        def _dt(v):
            return str(v) if v else ""

        def _payments(e):
            # Same "3 ACC · 2 PDNG" rendering as the operational table's pills:
            # {status: count} sorted by count desc then status.
            agg = getattr(e, "payment_statuses", None)
            if not agg:
                return ""
            return " · ".join(
                f"{count} {status}"
                for status, count in sorted(agg.items(), key=lambda kv: (-kv[1], kv[0]))
            )

        def _payment_timestamp(e):
            # Same span rendering as the operational table: a single timestamp
            # when every payment of the reco shares one, "min → max" otherwise.
            low = getattr(e, "payment_timestamp_min", None)
            high = getattr(e, "payment_timestamp_max", None)
            if low is None and high is None:
                return ""
            if low == high or high is None:
                return str(low)
            return f"{low} → {high}" if low is not None else str(high)

        # (header, value getter) pairs — one source of truth for the column order,
        # so the data rows can never drift from the header row.
        columns = [
            ("Flow", lambda e: flow_codes.get(e.flow_id, f"#{e.flow_id}")),
            ("Reco ID", lambda e: e.reco_id),
            ("Account", lambda e: e.account),
            ("Currency", lambda e: e.currency),
            ("Amount", lambda e: float(e.amount) if e.amount is not None else None),
            ("Direction", lambda e: _enum(e.direction)),
            ("Value date", lambda e: _dt(e.value_date)),
            ("Operation date", lambda e: _dt(e.operation_date)),
            ("Event type", lambda e: e.event_type),
            ("Transaction ID", lambda e: e.transaction_id),
            ("External ref", lambda e: e.external_ref),
            ("File name", lambda e: e.file_name),
            ("Particulars", lambda e: e.transaction_particulars),
            ("Ref no", lambda e: e.ref_no),
            ("Remarks 1", lambda e: e.remarks_1),
            ("Status", lambda e: _enum(e.status)),
            ("Matched at", lambda e: _dt(e.matched_at)),
            ("Payments", _payments),
            ("PaymentTimestamp", _payment_timestamp),
        ]

        for col_idx, (header, _getter) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_text
            cell.fill = header_fill

        for row_idx, e in enumerate(items, 2):
            for col_idx, (_header, getter) in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=getter(e))

        # Auto-width columns (cap at 40), mirroring the archive export.
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def list_match_groups(self, db: Session, *, flow_id=None, skip: int = 0, limit: int = 100) -> list:
        return match_group_repository.list_filtered(db, flow_id=flow_id, skip=skip, limit=limit)

    # --------------------------------------------------------------
    # File-source reco_id resolution (e.g. MT940 keys → finacle reco_id)
    # --------------------------------------------------------------
    def resolve_pending_references(self, db: Session) -> int:
        """Global safety-net sweep run before matching: fill reco_id for live PENDING
        entries from lookup-enabled file sources (e.g. MT940 +21 keys) from a matching
        finacle entry. Delegates to the repository, which is also used per-flow as the
        first step of file ingestion. Idempotent; returns the number resolved."""
        return reconciliation_entry_repository.resolve_flow_references(db)

    # --------------------------------------------------------------
    # Automatic engine
    # --------------------------------------------------------------
    def run_auto(self, db: Session, *, triggered_by: Optional[str] = None) -> ReconciliationRun:
        """Automatic engine. Single-writer, enforced by a Postgres advisory lock.

        Two concurrent run_auto deadlock each other: mark_matched walks
        reconciliation_entry through ix_reconciliation_entry_flow_status_date (so
        in value_date order) while the other run's move_matched_to_emargement
        walks the very same rows through ix_reconciliation_entry_reco_id —
        opposite lock orders over one set of tuples. Rather than order every
        write, the engine is made single-writer: it is a batch, not a query path,
        and a second concurrent run has no work to do anyway (the first one has
        already consumed the balanced groups).

        The lock is taken BEFORE creating the ReconciliationRun, so that a
        skipped run leaves no phantom finished_at=NULL row behind.
        """
        with try_advisory_lock(LockKey.RECONCILIATION_ENTRY_WRITER) as acquired:
            if not acquired:
                logger.info("reconciliation run skipped: another run holds the lock")
                raise ReconciliationAlreadyRunning(
                    "a reconciliation run is already in progress"
                )
            return self._run_auto_locked(db, triggered_by=triggered_by)

    def _run_auto_locked(self, db: Session, *, triggered_by: Optional[str]) -> ReconciliationRun:
        """Engine body. Only ever called holding RECONCILIATION_ENTRY_WRITER."""
        started = datetime.now(timezone.utc)
        run = reconciliation_run_repository.create(
            db,
            run=ReconciliationRun(started_at=started, triggered_by=triggered_by or "manual"),
        )
        # Captured now: a rollback expires the instance, and touching run.id then
        # fires a fresh SELECT — not something to do from an error handler.
        run_id = run.id
        groups_created = 0
        entries_matched = 0
        try:
            # Resolve reco_id for file entries (e.g. MT940) that reference a finacle
            # movement ingested earlier — so they can match in this run.
            self.resolve_pending_references(db)
            balanced = reconciliation_entry_repository.find_balanced_groups(db)
            for flow_id, reco_id, currency, total in balanced:
                mg = match_group_repository.create(
                    db,
                    mg=MatchGroup(
                        flow_id=flow_id,
                        reco_id=reco_id,
                        currency=currency,
                        total=total,
                        mode=MatchMode.AUTO,
                        created_at=datetime.now(timezone.utc),
                        reconciliation_run_id=run.id,
                    ),
                )
                count = reconciliation_entry_repository.mark_matched(
                    db,
                    flow_id=flow_id,
                    reco_id=reco_id,
                    currency=currency,
                    match_group_id=mg.id,
                )
                # Move matched entries from live to émargement table
                reconciliation_entry_repository.move_matched_to_emargement(
                    db,
                    flow_id=flow_id,
                    reco_id=reco_id,
                    currency=currency,
                )
                groups_created += 1
                entries_matched += count
            # Second reconciliation: per claim group, do the split parents add
            # up to the ghosts standing in for them? Tags/clears
            # movement_lot.parent_mismatch — after the matching, so a lot that
            # just matched still gets tagged when its parent side does not add
            # up (matching the counterpart is NOT validating the parents).
            movement_split_repository.refresh_parent_mismatch(db)
            db.commit()
        except Exception:
            # The failing statement (a deadlock victim, say) leaves the psycopg2
            # transaction aborted: without this rollback every later statement
            # raises InFailedSqlTransaction — including the bookkeeping below,
            # which would then mask the real error and leave the run unfinalized
            # (finished_at NULL) for good.
            db.rollback()
            raise
        finally:
            finished = datetime.now(timezone.utc)
            try:
                reconciliation_run_repository.update(
                    db,
                    run=run,
                    finished_at=finished,
                    groups_created=groups_created,
                    entries_matched=entries_matched,
                    duration_ms=int((finished - started).total_seconds() * 1000),
                )
            except Exception:
                # An exception raised in a finally REPLACES the one propagating.
                # Bookkeeping must never erase the root cause on the way out.
                logger.exception("failed to finalize reconciliation run %s", run_id)
        return run

    # --------------------------------------------------------------
    # Manual forcing
    # --------------------------------------------------------------
    def force_match(
        self,
        db: Session,
        *,
        entry_ids: List[int],
        comment: Optional[str],
        user_id: Optional[int],
    ) -> MatchGroup:
        """Force a manual match. Pre-flight controls per spec:
        - all entries must belong to the same flow
        - same currency
        - sum must be zero

        Built for baskets: the caller may hand over hundreds of ids spanning
        several reco_ids, assembled over several searches. That makes three
        things matter that did not when this only ever saw two hand-picked rows —
        the ids are read in one query, every bad id is reported at once rather
        than just the first, and a group that could not be marked in full is
        rolled back instead of being left half-built.
        """
        if not entry_ids:
            raise ValueError("at least one entry is required")

        # Deduplicate, preserving order: the same id sent twice would otherwise
        # be counted twice in the total and "balance" a group that does not.
        seen: set = set()
        ids = [i for i in entry_ids if not (i in seen or seen.add(i))]

        # One query for the whole basket, not one per id.
        found = reconciliation_entry_repository.get_many(db, entry_ids=ids)
        by_id = {e.id: e for e in found}

        # get_many reads the live table only, which is where PENDING lives; an id
        # missing from it is either gone or already émargé.
        missing = [i for i in ids if i not in by_id]
        if missing:
            raise ValueError(
                "entries not found or already reconciled: "
                + ", ".join(str(i) for i in missing)
            )
        not_pending = [e for e in found if e.status != EntryStatus.PENDING]
        if not_pending:
            raise ValueError(
                "entries are not pending: "
                + ", ".join(
                    f"{e.id} (status={getattr(e.status, 'value', e.status)})"
                    for e in not_pending
                )
            )

        entries = [by_id[i] for i in ids]

        flow_ids = {e.flow_id for e in entries}
        if len(flow_ids) != 1:
            raise ValueError("all entries must belong to the same flow")
        currencies = {e.currency for e in entries}
        if len(currencies) != 1:
            raise ValueError("all entries must share the same currency")
        total = sum((e.amount for e in entries), Decimal("0"))
        if total != Decimal("0"):
            raise ValueError("sum is not zero — manual match requires a balanced group (total = 0)")

        flow_id = entries[0].flow_id
        currency = entries[0].currency
        # Label the group with a reco_id only when the basket is unambiguous.
        # A basket deliberately spans several recos — picking the first non-null
        # one, as this used to, files the group under an arbitrary member.
        distinct_recos = {e.reco_id for e in entries if e.reco_id}
        reco_id = next(iter(distinct_recos)) if len(distinct_recos) == 1 else None

        mg = match_group_repository.create(
            db,
            mg=MatchGroup(
                flow_id=flow_id,
                reco_id=reco_id,
                currency=currency,
                total=total,
                mode=MatchMode.FORCED,
                created_by_user_id=user_id,
                created_at=datetime.now(timezone.utc),
                comment=comment,
            ),
        )
        marked = reconciliation_entry_repository.mark_forced(
            db, entry_ids=ids, match_group_id=mg.id
        )
        if marked != len(ids):
            # An entry left PENDING between the check above and the update — the
            # group would be short a leg and no longer sum to zero. Put the rows
            # back and drop the group rather than émarger a broken match.
            reconciliation_entry_repository.revert_forced(db, match_group_id=mg.id)
            match_group_repository.delete(db, match_group_id=mg.id)
            raise ValueError(
                f"only {marked} of {len(ids)} entries could be forced — some were "
                "reconciled concurrently; refresh and try again"
            )
        # Move forced entries to émargement
        reconciliation_entry_repository.move_to_emargement(db, entry_ids=ids)
        return mg

    # --------------------------------------------------------------
    # Manual exclusion
    # --------------------------------------------------------------
    def exclude(
        self,
        db: Session,
        *,
        entry_id: int,
        reason: str,
        user_id: Optional[int],
    ) -> Exclusion:
        if not reason or not reason.strip():
            raise ValueError("exclusion reason is mandatory")
        e = reconciliation_entry_repository.get_one(db, entry_id=entry_id)
        if e is None:
            raise ValueError(f"entry {entry_id} not found")
        if e.status == EntryStatus.MATCHED:
            raise ValueError("cannot exclude an already matched entry")

        excl = exclusion_repository.create(
            db,
            exclusion=Exclusion(
                entry_id=entry_id,
                entry_value_date=e.value_date,
                reason=reason.strip(),
                created_by_user_id=user_id,
                created_at=datetime.now(timezone.utc),
            ),
        )
        reconciliation_entry_repository.mark_excluded(db, entry_id=entry_id)
        # Move excluded entry to émargement
        reconciliation_entry_repository.move_to_emargement(db, entry_ids=[entry_id])
        return excl

    # --------------------------------------------------------------
    # Unexclude (reverse an exclusion)
    # --------------------------------------------------------------
    def unexclude(
        self,
        db: Session,
        *,
        entry_id: int,
        reason: str,
        user_id: Optional[int],
    ) -> Exclusion:
        if not reason or not reason.strip():
            raise ValueError("unexclude reason is mandatory")

        # Entry must exist in émargement with EXCLUDED status
        entry = reconciliation_entry_repository.get_emargement_entry(db, entry_id=entry_id)
        if entry is None:
            raise ValueError(f"entry {entry_id} not found in émargement table")
        if entry.status != EntryStatus.EXCLUDED:
            raise ValueError(f"entry {entry_id} is not excluded (status={entry.status})")

        # Active (non-cancelled) exclusion, if any. It may be None when a previous
        # partial unexclude already cancelled it but failed to move the entry — we
        # still recover by moving the (still EXCLUDED) entry back.
        excl = exclusion_repository.get_active_for_entry(db, entry_id=entry_id)
        records = exclusion_repository.list_for_entry(db, entry_id=entry_id)
        if excl is None and not records:
            raise ValueError(f"no exclusion found for entry {entry_id}")

        # Atomic: move the entry back to live AND cancel the exclusion, commit once.
        try:
            moved = reconciliation_entry_repository.move_from_emargement_to_live(
                db, entry_id=entry_id, commit=False
            )
            if moved == 0:
                db.rollback()
                raise ValueError(
                    f"failed to move entry {entry_id} back to live table "
                    "(it may already exist in the live table)"
                )
            if excl is not None:
                exclusion_repository.cancel(db, exclusion=excl, user_id=user_id, commit=False)
            db.commit()
        except Exception:
            db.rollback()
            raise

        return excl or records[0]


reconciliation_service = ReconciliationService()
