"""Archive service — snapshot of reconciliation state at a given date + daily activity."""
from __future__ import annotations

import io
from datetime import date, datetime, time, timezone
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, text
from sqlalchemy.orm import Session

from app.models.audit_log import UIActionLog
from app.models.exclusion import Exclusion
from app.models.reconciliation_entry import (
    EntryStatus,
    ReconciliationEntry,
    ReconciliationEntryEmargement,
)
from app.models.user import User
from app.schemas.time_machine import (
    DailyActivityEntry,
    DailyActivityResponse,
    SnapshotEntryResponse,
    SnapshotResponse,
    SnapshotSummary,
)


def _end_of_day(d: date) -> datetime:
    """Return the last instant of the given day (UTC)."""
    return datetime.combine(d, time(23, 59, 59, 999999), tzinfo=timezone.utc)


class ArchiveService:
    # ------------------------------------------------------------------
    # Snapshot — state of reconciliation at a given date
    # ------------------------------------------------------------------
    def get_snapshot(
        self,
        db: Session,
        *,
        snapshot_date: date,
        flow_id: Optional[int] = None,
        status: Optional[str] = None,
        skip: int = 0,
        limit: int = 200,
    ) -> SnapshotResponse:
        cutoff = _end_of_day(snapshot_date)

        # Use raw SQL for the UNION query across live and émargement tables.
        # Entries that existed at `cutoff`:
        #  1. Live table entries created before cutoff → they were PENDING at cutoff
        #  2. Émargement entries created before cutoff AND emarged before cutoff → final status
        #  3. Émargement entries created before cutoff AND emarged after cutoff → PENDING at cutoff
        #
        # For unexcluded entries (currently in live after unexclude):
        #  Check the exclusion table for active exclusions at cutoff.
        sql_parts = """
            SELECT id, flow_id, reco_id, account, currency, amount, direction,
                   value_date, operation_date, event_type, external_ref, file_name,
                   status_at_snapshot, match_group_id, matched_at, exclusion_reason
            FROM (
                -- Live table: entries created before cutoff
                SELECT e.id, e.flow_id, e.reco_id, e.account, e.currency, e.amount,
                       e.direction::text AS direction,
                       e.value_date, e.operation_date, e.event_type, e.external_ref, e.file_name,
                       CASE
                           WHEN ex.id IS NOT NULL THEN 'excluded'
                           WHEN LOWER(e.status::text) IN ('matched', 'forced') AND e.matched_at IS NOT NULL AND e.matched_at <= :cutoff
                               THEN LOWER(e.status::text)
                           ELSE 'pending'
                       END AS status_at_snapshot,
                       CASE
                           WHEN LOWER(e.status::text) IN ('matched', 'forced') AND e.matched_at IS NOT NULL AND e.matched_at <= :cutoff
                               THEN e.match_group_id
                           ELSE NULL
                       END AS match_group_id,
                       CASE
                           WHEN LOWER(e.status::text) IN ('matched', 'forced') AND e.matched_at IS NOT NULL AND e.matched_at <= :cutoff
                               THEN e.matched_at
                           ELSE NULL
                       END AS matched_at,
                       ex.reason AS exclusion_reason
                FROM reco.reconciliation_entry e
                LEFT JOIN reco.exclusion ex
                    ON ex.entry_id = e.id
                    AND ex.created_at <= :cutoff
                    AND (ex.cancelled_at IS NULL OR ex.cancelled_at > :cutoff)
                WHERE e.created_at <= :cutoff

                UNION ALL

                -- Émargement: entries emarged BEFORE cutoff → show with their final status
                SELECT em.id, em.flow_id, em.reco_id, em.account, em.currency, em.amount,
                       em.direction::text AS direction,
                       em.value_date, em.operation_date, em.event_type, em.external_ref, em.file_name,
                       LOWER(em.status::text) AS status_at_snapshot,
                       em.match_group_id, em.matched_at,
                       ex.reason AS exclusion_reason
                FROM reco.reconciliation_entry_emargement em
                LEFT JOIN reco.exclusion ex
                    ON ex.entry_id = em.id
                    AND ex.created_at <= :cutoff
                    AND (ex.cancelled_at IS NULL OR ex.cancelled_at > :cutoff)
                WHERE em.created_at <= :cutoff
                  AND em.emarged_at <= :cutoff

                UNION ALL

                -- Émargement: entries emarged AFTER cutoff → were still PENDING at cutoff
                SELECT em.id, em.flow_id, em.reco_id, em.account, em.currency, em.amount,
                       em.direction::text AS direction,
                       em.value_date, em.operation_date, em.event_type, em.external_ref, em.file_name,
                       'pending' AS status_at_snapshot,
                       NULL AS match_group_id, NULL AS matched_at,
                       NULL AS exclusion_reason
                FROM reco.reconciliation_entry_emargement em
                WHERE em.created_at <= :cutoff
                  AND em.emarged_at > :cutoff
            ) AS snapshot
        """

        params: dict = {"cutoff": cutoff}

        # Apply filters
        where_clauses = []
        if flow_id is not None:
            where_clauses.append("flow_id = :flow_id")
            params["flow_id"] = flow_id
        if status:
            where_clauses.append("status_at_snapshot = :status")
            params["status"] = status

        where_sql = ""
        if where_clauses:
            where_sql = " WHERE " + " AND ".join(where_clauses)

        # Count query
        count_sql = f"SELECT COUNT(*) FROM ({sql_parts}{where_sql}) AS cnt"
        total_count = db.execute(text(count_sql), params).scalar() or 0

        # Summary query (counts per status)
        summary_sql = f"""
            SELECT status_at_snapshot, COUNT(*) AS cnt
            FROM ({sql_parts}) AS summary_q
            {f"WHERE flow_id = :flow_id" if flow_id is not None else ""}
            GROUP BY status_at_snapshot
        """
        summary_params = {"cutoff": cutoff}
        if flow_id is not None:
            summary_params["flow_id"] = flow_id
        summary_rows = db.execute(text(summary_sql), summary_params).all()

        counts = {row.status_at_snapshot: row.cnt for row in summary_rows}
        summary = SnapshotSummary(
            snapshot_date=snapshot_date,
            total_entries=sum(counts.values()),
            pending_count=counts.get("pending", 0),
            matched_count=counts.get("matched", 0),
            forced_count=counts.get("forced", 0),
            excluded_count=counts.get("excluded", 0),
        )

        # Data query with pagination
        data_sql = f"""
            {sql_parts}{where_sql}
            ORDER BY value_date DESC, id DESC
            OFFSET :skip LIMIT :limit
        """
        params["skip"] = skip
        params["limit"] = limit
        rows = db.execute(text(data_sql), params).all()

        items = [
            SnapshotEntryResponse(
                id=r.id,
                flow_id=r.flow_id,
                reco_id=r.reco_id,
                account=r.account,
                currency=r.currency,
                amount=r.amount,
                direction=r.direction,
                value_date=r.value_date,
                operation_date=r.operation_date,
                event_type=r.event_type,
                external_ref=r.external_ref,
                file_name=r.file_name,
                status_at_snapshot=r.status_at_snapshot,
                match_group_id=r.match_group_id,
                matched_at=r.matched_at,
                exclusion_reason=r.exclusion_reason,
            )
            for r in rows
        ]

        return SnapshotResponse(summary=summary, items=items, total_count=total_count)

    # ------------------------------------------------------------------
    # Daily activity — UI actions for a specific date
    # ------------------------------------------------------------------
    def get_daily_activity(
        self,
        db: Session,
        *,
        activity_date: date,
        flow_id: Optional[int] = None,
        user_id: Optional[int] = None,
        skip: int = 0,
        limit: int = 200,
    ) -> DailyActivityResponse:
        start = datetime.combine(activity_date, time(0, 0, 0), tzinfo=timezone.utc)
        end = _end_of_day(activity_date)

        q = (
            db.query(UIActionLog, User.full_name)
            .outerjoin(User, UIActionLog.user_id == User.id)
            .filter(UIActionLog.ts >= start, UIActionLog.ts <= end)
        )
        if user_id is not None:
            q = q.filter(UIActionLog.user_id == user_id)

        total = q.count()
        rows = q.order_by(UIActionLog.ts.desc()).offset(skip).limit(limit).all()

        actions = [
            DailyActivityEntry(
                id=log.id,
                user_id=log.user_id,
                user_name=full_name,
                action=log.action,
                target_type=log.target_type,
                target_id=log.target_id,
                details=log.details,
                ts=log.ts,
            )
            for log, full_name in rows
        ]

        return DailyActivityResponse(
            activity_date=activity_date,
            actions=actions,
            total_count=total,
        )

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------
    def export_snapshot_excel(
        self,
        db: Session,
        *,
        snapshot_date: date,
        flow_id: Optional[int] = None,
    ) -> io.BytesIO:
        """Generate an Excel file with the snapshot data."""
        # Fetch all entries (no pagination)
        snapshot = self.get_snapshot(
            db,
            snapshot_date=snapshot_date,
            flow_id=flow_id,
            skip=0,
            limit=100_000,
        )

        wb = Workbook()

        # -- Summary sheet --
        ws_summary = wb.active
        ws_summary.title = "Summary"
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="2B2D42", end_color="2B2D42", fill_type="solid")
        header_text = Font(bold=True, color="FFFFFF")

        ws_summary["A1"] = "Archive Snapshot"
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary["A3"] = "Date"
        ws_summary["B3"] = str(snapshot.summary.snapshot_date)
        ws_summary["A4"] = "Total entries"
        ws_summary["B4"] = snapshot.summary.total_entries
        ws_summary["A5"] = "Pending"
        ws_summary["B5"] = snapshot.summary.pending_count
        ws_summary["A6"] = "Matched"
        ws_summary["B6"] = snapshot.summary.matched_count
        ws_summary["A7"] = "Forced"
        ws_summary["B7"] = snapshot.summary.forced_count
        ws_summary["A8"] = "Excluded"
        ws_summary["B8"] = snapshot.summary.excluded_count

        for row in range(3, 9):
            ws_summary.cell(row=row, column=1).font = Font(bold=True)

        ws_summary.column_dimensions["A"].width = 18
        ws_summary.column_dimensions["B"].width = 20

        # -- Data sheet --
        ws_data = wb.create_sheet("Entries")
        headers = [
            "ID", "Flow ID", "Reco ID", "Account", "Currency", "Amount",
            "Direction", "Value Date", "Operation Date", "Event Type",
            "External Ref", "File Name", "Status", "Match Group ID",
            "Matched At", "Exclusion Reason",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = header_text
            cell.fill = header_fill

        for row_idx, entry in enumerate(snapshot.items, 2):
            ws_data.cell(row=row_idx, column=1, value=entry.id)
            ws_data.cell(row=row_idx, column=2, value=entry.flow_id)
            ws_data.cell(row=row_idx, column=3, value=entry.reco_id)
            ws_data.cell(row=row_idx, column=4, value=entry.account)
            ws_data.cell(row=row_idx, column=5, value=entry.currency)
            ws_data.cell(row=row_idx, column=6, value=float(entry.amount))
            ws_data.cell(row=row_idx, column=7, value=entry.direction)
            ws_data.cell(row=row_idx, column=8, value=str(entry.value_date) if entry.value_date else "")
            ws_data.cell(row=row_idx, column=9, value=str(entry.operation_date) if entry.operation_date else "")
            ws_data.cell(row=row_idx, column=10, value=entry.event_type)
            ws_data.cell(row=row_idx, column=11, value=entry.external_ref)
            ws_data.cell(row=row_idx, column=12, value=entry.file_name)
            ws_data.cell(row=row_idx, column=13, value=entry.status_at_snapshot)
            ws_data.cell(row=row_idx, column=14, value=entry.match_group_id)
            ws_data.cell(row=row_idx, column=15, value=str(entry.matched_at) if entry.matched_at else "")
            ws_data.cell(row=row_idx, column=16, value=entry.exclusion_reason)

        # Auto-width columns
        for col in ws_data.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_data.column_dimensions[col_letter].width = min(max_length + 2, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output


archive_service = ArchiveService()
