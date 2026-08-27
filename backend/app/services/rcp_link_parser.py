"""Pure parsing + control of the RCP link extracts (no DB, no app imports).

Two file families land here:

* the LINK workbook (``SP_LINK_REPORT_*.xlsx``) — one row per Paysis movement,
  carrying ``pmsgid``/``msgid``, ``num_0f_records`` (sic — the extract ships the typo) and
  ``settlementamt``. Only the ``RCP`` rows matter: they are the return/reject
  batches booked on the float, the ones the datamart cannot link back to the
  bulk they undo. ``NCP`` rows are ordinary payment batches and are skipped.
* N PAYMENT DUMPS — extracts of the return/reject table. Their shape varies
  (the return dumps are camelCase and carry three amounts; the reject dump is
  lowercase and only carries ``orgnlsttlmamt``), so every header is folded to
  lowercase and the amount is taken from the first column that exists — the
  ORIGINAL settlement amount first, see ``DUMP_AMOUNT``. The message id moved
  from ``msgid`` to ``pmsgid`` on 2026-08-14 and both are accepted, ``pmsgid``
  winning when a file carries both (the recall dump does). A dump WITHOUT a
  msgid column cannot be matched at all — that is a fact about the extract,
  reported per file rather than silently ignored.

The extracts are CUMULATIVE, not deltas: every daily export re-ships the whole
history, filling in what the previous one left empty. So several days are read
together and ``build_dump_index`` deduplicates on ``entitySrlNum``; the newest
link workbook is a strict superset of the older ones and is used alone.

The control is the conservation proof of the whole reattribution: a movement's
dump rows must be exactly ``num_0f_records`` and must add up to
``settlementamt``. When they do, splitting the movement over its target lots is
exact to the cent; when they do not, the movement is still proposed (decision:
non-blocking) and the gap will be tagged by the existing claim-group
reconciliation once committed.
"""
from __future__ import annotations

import csv
import hashlib
import io
from dataclasses import dataclass, field
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

# ── xlsx columns (headers folded to lowercase) ───────────────────────
LINK_MSGID = ("pmsgid", "msgid", "messageid", "message_id")
LINK_SERVICE = ("serviceid", "service_id")
LINK_DIRECTION = ("direction",)
# The shipped extract spells it ``num_0f_records`` (zero, not 'o'); accept both.
LINK_COUNT = ("num_0f_records", "num_of_records", "numofrecords", "nb_records")
LINK_AMOUNT = ("settlementamt", "settlementamount", "sttlmamt")
LINK_SPDATE = ("spdate",)
LINK_TRANDATE = ("trandate",)
LINK_TRANID = ("tranid", "tran_id")
LINK_PAYSYS = ("paysysid", "paysys_id")

# ── dump columns ─────────────────────────────────────────────────────
DUMP_MSGID = ("pmsgid", "msgid", "messageid", "message_id")
DUMP_SRL = ("entitysrlnum", "entity_srl_num")
DUMP_ORIG = ("origentityid", "orig_entity_id", "originalentityid")
# First one present wins, and the ORIGINAL settlement amount comes first: it is
# what the float booked, which is what the ghosts must add up to. The returned
# amount is not the same thing — on an inward SDD return it carries the
# interbank return fee on top (measured on the 2026-08-20 extract, entitySrlNum
# 000009216272: returnSttlmAmt 77.00, orgnlSttlmAmt 74.00, and the link workbook
# books 74.00). 151 rows diverge over six days, and taking the bigger one made
# the ghosts exceed the booking, so ``validate_slices`` refused ~37 movements
# outright. The reject dumps only ship ``orgnlsttlmamt``; the recall dump ships
# neither, hence its own columns.
DUMP_AMOUNT = (
    "orgnlsttlmamt",
    "orgnlintrbksttlmamt",
    "recalledinstdamt",
    "recallsttlmamt",
    "returnsttlmamt",
    "instructedamt",
    "sttlmamt",
)
DUMP_SERVICE = ("serviceid", "service_id")
DUMP_TYPE = ("servicetype", "service_type")
DUMP_STATUS = ("status",)
DUMP_REASON = ("rtrrsncd", "rsncd", "reasoncode")
DUMP_REASON_DESC = ("rsndescr", "addlinfo", "reasondescription")
DUMP_DATE = ("intrbnksttlmdate", "settlementdate", "_createdon", "createdon")

# The aggregated return/reject batches worth reattributing. NCP (ordinary
# payment batches) is deliberately absent: those movements settle normally.
REATTRIBUTABLE_SERVICES = frozenset({"RCP", "RCC", "RRS", "WCC"})

CENT = Decimal("0.01")

# Control statuses.
CTRL_OK = "OK"
CTRL_NOT_FOUND = "NOT_FOUND"
CTRL_COUNT_MISMATCH = "COUNT_MISMATCH"
CTRL_AMOUNT_MISMATCH = "AMOUNT_MISMATCH"
CTRL_DUPLICATE_MOVEMENT = "DUPLICATE_MOVEMENT"
# The movement books 0.00 while its rows carry value: Finacle settles an outward
# reject batch a day or two AFTER listing it (measured: 9 rows of the workbook
# went from 0.00 to their real amount over the six daily extracts). Nothing to
# arbitrate — the next extract fixes it by itself.
CTRL_NOT_SETTLED = "NOT_SETTLED"

# Bucket kinds — must stay identical to movement_lot.BUCKET_* and to the DAG's
# ``payment_bucket``; a lot is looked up by these very values.
BUCKET_PAIR = "PAIR"
BUCKET_PACS_ONLY = "PACS_ONLY"
BUCKET_MSGID_ONLY = "MSGID_ONLY"
BUCKET_PO = "PO"


def q2(value: Decimal) -> Decimal:
    return value.quantize(CENT, rounding=ROUND_HALF_UP)


def parse_decimal(raw: Any) -> Optional[Decimal]:
    """Amounts arrive as floats (xlsx) or strings ('70.00', '1 234,56')."""
    if raw is None:
        return None
    if isinstance(raw, Decimal):
        return raw
    if isinstance(raw, (int, float)):
        return Decimal(str(raw))
    s = str(raw).strip().replace(" ", "").replace(" ", "").replace(" ", "")
    if not s:
        return None
    if "," in s and "." in s:
        # Whichever separator comes last is the decimal one.
        s = s.replace(",", "") if s.rfind(".") > s.rfind(",") else s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def parse_int(raw: Any) -> Optional[int]:
    dec = parse_decimal(raw)
    if dec is None:
        return None
    try:
        return int(dec)
    except (ValueError, ArithmeticError):
        return None


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _pick(row: Dict[str, Any], names: Sequence[str]) -> Any:
    for name in names:
        if name in row:
            return row[name]
    return None


def _has_any(headers: Iterable[str], names: Sequence[str]) -> bool:
    header_set = set(headers)
    return any(name in header_set for name in names)


# ---------------------------------------------------------------------------
# Value objects
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class LinkMovement:
    """One reattributable row of the link workbook (RCP/RCC/RRS/WCC)."""
    msgid: str
    service_id: str
    direction: str
    num_records: int
    settlement_amount: Decimal
    sp_date: str = ""
    tran_date: str = ""
    tran_id: str = ""
    paysys_id: str = ""
    row_number: int = 0


@dataclass(frozen=True)
class DumpRow:
    """One return/reject row, whatever dump it came from."""
    msgid: str
    entity_srl_num: str
    orig_entity_id: str
    amount: Decimal
    service_id: str = ""
    service_type: str = ""
    status: str = ""
    reason_code: str = ""
    reason_desc: str = ""
    settlement_date: str = ""
    file_name: str = ""


@dataclass
class FileReport:
    """What one uploaded file actually contained — the extract-quality signal."""
    file_name: str
    rows: int = 0
    rows_with_msgid: int = 0
    distinct_msgids: int = 0
    has_msgid_column: bool = True
    amount_column: str = ""
    delimiter: str = ""
    error: str = ""
    # {serviceid: rows} over the WHOLE workbook — kept types and skipped ones
    # alike. A service that never appears, or one nobody expected, shows up here
    # instead of silently changing what the run covers.
    services: Dict[str, int] = field(default_factory=dict)


@dataclass
class ControlResult:
    """Part 1: does the dump agree with the booked movement?"""
    msgid: str
    status: str
    expected_count: int
    found_count: int
    expected_amount: Decimal
    found_amount: Decimal
    files: List[str] = field(default_factory=list)
    service_id: str = ""
    direction: str = ""
    tran_id: str = ""
    sp_date: str = ""

    @property
    def delta_count(self) -> int:
        return self.found_count - self.expected_count

    @property
    def delta_amount(self) -> Decimal:
        return q2(self.found_amount - self.expected_amount)

    @property
    def ok(self) -> bool:
        return self.status == CTRL_OK


# ---------------------------------------------------------------------------
# Link workbook
# ---------------------------------------------------------------------------

def parse_link_workbook(
    content: bytes,
    file_name: str = "",
    services: Optional[Iterable[str]] = None,
) -> Tuple[List[LinkMovement], FileReport]:
    """The reattributable rows of the link workbook, plus what the file held.

    ``services`` defaults to ``REATTRIBUTABLE_SERVICES`` (RCP/RCC/RRS/WCC). Every
    row is counted per ``serviceid`` in the report — including the dropped NCP
    payment batches — so widening or narrowing the set is visible rather than
    silent, and a service type nobody expected shows up on the first run.
    """
    kept = {s.strip().upper() for s in (services or REATTRIBUTABLE_SERVICES)}
    from openpyxl import load_workbook  # local import: keeps the module import-cheap

    report = FileReport(file_name=file_name or "link.xlsx")
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # unreadable/not an xlsx
        report.error = f"unreadable workbook: {exc}"
        return [], report

    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            report.error = "empty worksheet"
            return [], report
        headers = [_clean(h).lower() for h in header_row]
        if not _has_any(headers, LINK_MSGID):
            report.error = "no msgid column"
            report.has_msgid_column = False
            return [], report

        movements: List[LinkMovement] = []
        for index, values in enumerate(rows, start=2):
            if values is None or all(v is None or _clean(v) == "" for v in values):
                continue
            report.rows += 1
            row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
            msgid = _clean(_pick(row, LINK_MSGID))
            if not msgid:
                continue
            report.rows_with_msgid += 1
            service = _clean(_pick(row, LINK_SERVICE)).upper()
            report.services[service or "?"] = report.services.get(service or "?", 0) + 1
            if service not in kept:
                continue
            count = parse_int(_pick(row, LINK_COUNT))
            amount = parse_decimal(_pick(row, LINK_AMOUNT))
            movements.append(
                LinkMovement(
                    msgid=msgid,
                    service_id=service,
                    direction=_clean(_pick(row, LINK_DIRECTION)).upper(),
                    num_records=count if count is not None else 0,
                    settlement_amount=q2(amount) if amount is not None else Decimal("0.00"),
                    sp_date=_clean(_pick(row, LINK_SPDATE)),
                    tran_date=_clean(_pick(row, LINK_TRANDATE)),
                    tran_id=_clean(_pick(row, LINK_TRANID)),
                    paysys_id=_clean(_pick(row, LINK_PAYSYS)),
                    row_number=index,
                )
            )
        report.distinct_msgids = len({m.msgid for m in movements})
        return movements, report
    finally:
        try:
            workbook.close()
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Payment dumps
# ---------------------------------------------------------------------------

def sniff_delimiter(header_line: str) -> str:
    """The separator the dump uses — ';' in every extract so far, but a CSV
    export can flip to ',' or a tab without warning."""
    counts = {sep: header_line.count(sep) for sep in (";", ",", "\t", "|")}
    best = max(counts, key=lambda sep: counts[sep])
    return best if counts[best] else ";"


# First line of a psql export that shipped ONE concatenated column per row
# instead of the columns themselves.
WRAPPED_EXPORT_HEADER = "?column?"


def unwrap_quoted_export(content: str) -> Tuple[str, bool]:
    """Undo a ``"?column?"`` export → (text, was_wrapped).

    ``extract/20260820/returnevent_sddxb_O.csv`` arrives as a first line
    ``"?column?"`` followed by rows whose entire ;-separated content sits inside
    one pair of quotes. Read as-is the file has no ``pmsgid`` column at all and
    its 1 633 rows disappear without a word. Unwrapping is a pure text step
    ahead of the CSV reader: drop the fake header, take one layer of quotes off
    each record, undouble the escaped ones. A record spanning several lines is
    accumulated until its quotes close, so an embedded newline cannot cut a row
    in two.
    """
    lines = content.splitlines()
    if not lines or lines[0].strip().strip('"') != WRAPPED_EXPORT_HEADER:
        return content, False

    records: List[str] = []
    buffer: Optional[str] = None
    for line in lines[1:]:
        buffer = line if buffer is None else f"{buffer}\n{line}"
        record = buffer.strip()
        if not record.startswith('"'):
            records.append(record)          # already bare
            buffer = None
        elif record.endswith('"') and len(record) >= 2:
            records.append(record[1:-1].replace('""', '"'))
            buffer = None
    if buffer is not None:                  # unterminated quote — keep the text
        records.append(buffer.strip().strip('"'))
    return "\n".join(records), True


def parse_dump_file(content: bytes, file_name: str) -> Tuple[List[DumpRow], FileReport]:
    """Rows of one return/reject dump, keeping only those carrying a msgid.

    A file with no ``msgid`` column at all yields nothing and says so: that is
    the 2026-08-12 ``reject_o`` case, where the column existed but was empty for
    every row — the reason 7 movements could not be completed.
    """
    report = FileReport(file_name=file_name)
    text = content.decode("utf-8-sig", errors="replace")
    if not text.strip():
        report.error = "empty file"
        return [], report

    text, wrapped = unwrap_quoted_export(text)
    if wrapped:
        # Reported, never silent: the extract shipped a broken export and the
        # operator should know the rows were recovered rather than read.
        report.error = "export psql enveloppé (\"?column?\") — déquoté"
    if not text.strip():
        report.error = (report.error + "; " if report.error else "") + "empty file"
        return [], report

    first_line = text.splitlines()[0]
    delimiter = sniff_delimiter(first_line)
    report.delimiter = delimiter
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = [_clean(h).lower() for h in (reader.fieldnames or [])]
    if not headers:
        report.error = "no header"
        return [], report
    if not _has_any(headers, DUMP_MSGID):
        report.has_msgid_column = False
        report.error = "no msgid column — cannot be matched"

    amount_column = next((name for name in DUMP_AMOUNT if name in headers), "")
    report.amount_column = amount_column
    if not amount_column:
        report.error = (report.error + "; " if report.error else "") + "no amount column"

    rows: List[DumpRow] = []
    seen_msgids = set()
    for raw in reader:
        report.rows += 1
        row = {headers[i]: raw.get(field)
               for i, field in enumerate(reader.fieldnames or []) if i < len(headers)}
        msgid = _clean(_pick(row, DUMP_MSGID))
        if not msgid:
            continue
        report.rows_with_msgid += 1
        seen_msgids.add(msgid)
        amount = parse_decimal(row.get(amount_column)) if amount_column else None
        rows.append(
            DumpRow(
                msgid=msgid,
                entity_srl_num=_clean(_pick(row, DUMP_SRL)),
                orig_entity_id=_clean(_pick(row, DUMP_ORIG)),
                amount=q2(amount) if amount is not None else Decimal("0.00"),
                service_id=_clean(_pick(row, DUMP_SERVICE)).upper(),
                service_type=_clean(_pick(row, DUMP_TYPE)).upper(),
                status=_clean(_pick(row, DUMP_STATUS)).upper(),
                reason_code=_clean(_pick(row, DUMP_REASON)),
                reason_desc=_clean(_pick(row, DUMP_REASON_DESC)),
                settlement_date=_clean(_pick(row, DUMP_DATE)),
                file_name=file_name,
            )
        )
    report.distinct_msgids = len(seen_msgids)
    return rows, report


def build_dump_index(rows: Iterable[DumpRow]) -> Tuple[Dict[str, List[DumpRow]], int]:
    """({msgid: rows} deduplicated ACROSS files, rows dropped as duplicates).

    The extracts overlap — each daily export re-ships the whole history, so the
    six folders of 2026-08 hold 207 202 rows for 40 519 distinct ones — and a
    duplicated row would inflate both the control count and a ghost's amount.
    ``entitySrlNum`` is the table's own key; when a dump omits it, the row's
    business content stands in.

    The count comes back so the run can SAY how much it dropped: silently
    collapsing 80% of the uploaded rows is exactly the kind of thing an operator
    must see confirmed rather than infer.
    """
    index: Dict[str, List[DumpRow]] = {}
    seen: set = set()
    duplicates = 0
    for row in rows:
        identity = (
            ("srl", row.entity_srl_num)
            if row.entity_srl_num
            else ("row", row.msgid, row.orig_entity_id, str(row.amount), row.reason_code)
        )
        if identity in seen:
            duplicates += 1
            continue
        seen.add(identity)
        index.setdefault(row.msgid, []).append(row)
    return index, duplicates


# ---------------------------------------------------------------------------
# Control (part 1)
# ---------------------------------------------------------------------------

def control_movement(movement: LinkMovement, rows: Sequence[DumpRow]) -> ControlResult:
    """Count + sum of a movement's dump rows against what it booked."""
    found_amount = q2(sum((r.amount for r in rows), Decimal("0")))
    result = ControlResult(
        msgid=movement.msgid,
        status=CTRL_OK,
        expected_count=movement.num_records,
        found_count=len(rows),
        expected_amount=q2(movement.settlement_amount),
        found_amount=found_amount,
        files=sorted({r.file_name for r in rows}),
        service_id=movement.service_id,
        direction=movement.direction,
        tran_id=movement.tran_id,
        sp_date=movement.sp_date,
    )
    if not rows:
        result.status = CTRL_NOT_FOUND
    elif not movement.settlement_amount and found_amount:
        # Listed but not settled yet — the workbook fills the amount in a day or
        # two later. Splitting it now would price the ghosts above a 0.00
        # booking, which the commit refuses anyway.
        result.status = CTRL_NOT_SETTLED
    elif len(rows) != movement.num_records:
        result.status = CTRL_COUNT_MISMATCH
    elif found_amount != q2(movement.settlement_amount):
        result.status = CTRL_AMOUNT_MISMATCH
    return result


def control_all(
    movements: Sequence[LinkMovement], index: Dict[str, List[DumpRow]]
) -> Tuple[List[ControlResult], List[str]]:
    """Control every RCP movement; also flags msgids the workbook lists twice
    (two movements claiming the same dump rows would each take them whole)."""
    counts: Dict[str, int] = {}
    for movement in movements:
        counts[movement.msgid] = counts.get(movement.msgid, 0) + 1
    duplicates = sorted(msgid for msgid, n in counts.items() if n > 1)

    results = []
    for movement in movements:
        result = control_movement(movement, index.get(movement.msgid, []))
        if movement.msgid in counts and counts[movement.msgid] > 1:
            result.status = CTRL_DUPLICATE_MOVEMENT
        results.append(result)
    return results, duplicates


# ---------------------------------------------------------------------------
# Buckets & ghost identity (mirrors shared/dags/reco_datamart_bb.py)
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class BucketKey:
    """A lot's identity. Components UPPERCASED, absent ones '' — exactly like
    the DAG's ``BucketKey``, because the lot row is looked up on these values
    and the datamart's collation is case-insensitive."""
    kind: str
    pacs: str = ""
    msgid: str = ""
    po: str = ""
    ref: str = ""

    def __post_init__(self) -> None:
        for name in ("pacs", "msgid", "po", "ref"):
            object.__setattr__(self, name, (getattr(self, name) or "").upper())

    def label(self) -> str:
        parts = [p for p in (self.pacs, self.msgid, self.po, self.ref) if p]
        return f"{self.kind}:{'|'.join(parts)}" if parts else self.kind


def payment_bucket(pacs: Optional[str], msgid: Optional[str], po: Optional[str]) -> Optional[BucketKey]:
    """The bucket one payment belongs to — copied from ``payment_bucket`` in
    shared/dags/reco_datamart_bb.py so a returned payment resolves to the very
    lot its original movement was filed under. PO takes precedence over MSGID:
    a payment with no pacs008 is a single, reached through its PaymentNumber.
    """
    pacs, msgid, po = _clean(pacs), _clean(msgid), _clean(po)
    if pacs and msgid:
        return BucketKey(BUCKET_PAIR, pacs=pacs, msgid=msgid)
    if pacs:
        return BucketKey(BUCKET_PACS_ONLY, pacs=pacs)
    if po:
        return BucketKey(BUCKET_PO, po=po)
    if msgid:
        return BucketKey(BUCKET_MSGID_ONLY, msgid=msgid)
    return None


def ghost_external_ref(claim: Tuple[str, str], key: BucketKey) -> str:
    """Identity of one reattribution ghost — a pure function of (claim, bucket),
    same shape as the DAG's ``group_ghost_ref``. No date, no parent reference:
    re-committing the same movement upserts the very same rows. Bounded to
    ``external_ref`` VARCHAR(128)."""
    claim_type, claim_value = claim
    claim_tag = hashlib.sha1(f"{claim_type}|{claim_value}".encode("utf-8")).hexdigest()[:8]
    bucket_raw = f"{key.kind}|{key.pacs}|{key.msgid}|{key.po}|{key.ref}"
    bucket_tag = hashlib.sha1(bucket_raw.encode("utf-8")).hexdigest()[:10]
    return f"KEY:{claim_value[:88]}~{claim_tag}~{bucket_tag}"
